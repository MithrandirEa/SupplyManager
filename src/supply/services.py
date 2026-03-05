"""
Service d'import d'articles depuis un fichier CSV ou Excel.
"""
import csv
from datetime import datetime

import openpyxl

from supplier.models import Supplier
from supply.models import Item, ItemsCategory

# Mapping des colonnes basé sur l'export
COL_NAME = 'Nom'
COL_CATEGORY = 'Catégorie'
COL_AVAILABLE = 'Disponible'
COL_OUTSIDE = 'Hors site'
COL_TOTAL = 'Total'
COL_EXCESS = 'Excédent'
COL_INFO = 'Informations'
COL_SUPPLIERS = 'Fournisseurs'
COL_STOCK_DATE = 'Date entrée stock'
COL_INV_QTY = 'Qté dernier inventaire'
COL_INV_DATE = 'Date dernier inventaire'


def _parse_int(val):
    """Parse une valeur en entier, retourne 0 si impossible."""
    if val is None or str(val).strip() == '':
        return 0
    try:
        return int(float(str(val).replace(',', '.')))
    except (ValueError, TypeError):
        return 0


def _read_csv(import_file):
    """Lit un fichier CSV et retourne une liste de dictionnaires."""
    content = import_file.read().decode('utf-8-sig')
    lines = [line for line in content.splitlines() if line.strip()]
    reader = csv.DictReader(lines, delimiter=';')
    return list(reader)


def _read_excel(import_file):
    """Lit un fichier Excel et retourne une liste de dictionnaires."""
    wb = openpyxl.load_workbook(import_file, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        has_data = False
        for i, cell_val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = cell_val
                if cell_val:
                    has_data = True
        if has_data:
            data_rows.append(row_dict)
    return data_rows


def _process_row(row, user, update_existing):
    """
    Traite une ligne de données d'import.

    Returns:
        'created' si un nouvel article a été créé,
        'updated' si un article existant a été mis à jour,
        None si la ligne a été ignorée.
    """
    name_val = row.get(COL_NAME)
    if not name_val:
        return None
    name = str(name_val).strip()

    # Catégorie
    cat_val = row.get(COL_CATEGORY)
    cat_name = str(cat_val).strip() if cat_val else 'Sans catégorie'
    category, _ = ItemsCategory.objects.get_or_create(name=cat_name)

    # Recherche Item
    item = Item.objects.filter(name=name, category=category).first()
    is_new = False

    if item:
        if not update_existing:
            return None
    else:
        item = Item(name=name, category=category)
        item.created_by = user
        is_new = True

    # Valeurs numériques
    item.available_quantity = _parse_int(row.get(COL_AVAILABLE))
    item.outside_quantity = _parse_int(row.get(COL_OUTSIDE))
    item.excess_quantity = _parse_int(row.get(COL_EXCESS))
    item.last_inventory_quantity = _parse_int(row.get(COL_INV_QTY))

    # Total
    total_val = _parse_int(row.get(COL_TOTAL))
    if total_val > 0:
        item.total_quantity = total_val
    else:
        item.total_quantity = item.available_quantity + item.outside_quantity

    # Info
    info_val = row.get(COL_INFO)
    item.information = str(info_val) if info_val else ''

    # Date entrée stock
    stock_date_val = row.get(COL_STOCK_DATE)
    if stock_date_val:
        if isinstance(stock_date_val, datetime):
            item.stock_entry_date = stock_date_val.date()
        elif isinstance(stock_date_val, str):
            try:
                item.stock_entry_date = datetime.strptime(
                    stock_date_val, '%d/%m/%Y'
                ).date()
            except (ValueError, TypeError):
                pass

    item.save()

    # Fournisseurs
    sup_val = row.get(COL_SUPPLIERS)
    if sup_val:
        sup_names = [s.strip() for s in str(sup_val).split(',') if s.strip()]
        for s_name in sup_names:
            supplier, _ = Supplier.objects.get_or_create(name=s_name)
            item.suppliers.add(supplier)

    return 'created' if is_new else 'updated'


def process_import(import_file, update_existing, user):
    """
    Traite l'import complet d'un fichier CSV ou Excel.

    Returns:
        tuple (created_count, updated_count, errors)
    """
    # Lecture du fichier
    if import_file.name.endswith('.csv'):
        data_rows = _read_csv(import_file)
    elif import_file.name.endswith(('.xlsx', '.xls')):
        data_rows = _read_excel(import_file)
    else:
        return 0, 0, ["Format de fichier non supporté."]

    created_count = 0
    updated_count = 0
    errors = []

    for idx, row in enumerate(data_rows):
        line_num = idx + 2
        try:
            result = _process_row(row, user, update_existing)
            if result == 'created':
                created_count += 1
            elif result == 'updated':
                updated_count += 1
        except Exception as e_row:
            errors.append(f"Ligne {line_num} : {e_row}")

    return created_count, updated_count, errors
