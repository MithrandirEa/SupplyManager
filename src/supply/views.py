import csv
import io
from datetime import datetime

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from django.urls import reverse

from authentication.decorators import role_required
from supplier.models import Supplier
from supply.models import Item, ItemsCategory

from .forms import (ChangeItemForm, CreateCategoryForm, CreateItemForm,
                    ImportItemsForm)


@role_required(['ADMIN', 'DIRECTOR'])
def create_category(request):
    if request.method == 'POST':
        form = CreateCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(f"{reverse('supplies_management')}#items")
    else:
        form = CreateCategoryForm()

    return render(request, 'create_category.html', {'form': form})


@role_required(['ADMIN', 'DIRECTOR'])
def create_item(request):
    if request.method == 'POST':
        form = CreateItemForm(request.POST)
        if form.is_valid():
            new_item = form.save(commit=False)
            new_item.available_quantity = (
                new_item.total_quantity - new_item.outside_quantity
            )
            new_item.created_by = request.user
            new_item.save()

            # Récupérer les fournisseurs sélectionnés depuis les inputs hidden
            selected_suppliers = request.POST.getlist('suppliers')
            if selected_suppliers:
                new_item.suppliers.set(selected_suppliers)

            return redirect(f"{reverse('supplies_management')}#items")
    else:
        form = CreateItemForm()

    all_suppliers = Supplier.objects.all().order_by('name')
    return render(request, 'create_item.html', {
        'form': form,
        'all_suppliers': all_suppliers,
        'selected_supplier_ids': []
    })


@role_required(['ADMIN', 'DIRECTOR'])
def change_item(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    if request.method == 'POST':
        form = ChangeItemForm(request.POST, instance=item)
        if form.is_valid():
            change_item = form.save(commit=False)
            change_item.available_quantity = (
                change_item.total_quantity - change_item.outside_quantity
            )
            change_item.save()

            # Récupérer les fournisseurs sélectionnés depuis les inputs hidden
            selected_suppliers = request.POST.getlist('suppliers')
            change_item.suppliers.set(selected_suppliers)

            return redirect(f"{reverse('supplies_management')}#items")
    else:
        form = ChangeItemForm(instance=item)

    all_suppliers = Supplier.objects.all().order_by('name')
    selected_supplier_ids = list(item.suppliers.values_list('id', flat=True))

    return render(request, 'change_item.html', {
        'form': form,
        'item': item,
        'all_suppliers': all_suppliers,
        'selected_supplier_ids': selected_supplier_ids
    })


@require_POST
@role_required(['ADMIN', 'DIRECTOR'])
def delete_item(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    item.delete()
    return redirect(f"{reverse('supplies_management')}#items")


@role_required(['ADMIN', 'DIRECTOR'])
def import_items(request):
    """
    Vue pour importer des articles depuis un fichier CSV ou Excel.
    """
    import csv
    # import io  # Pas utilisé directement si on passe par openpyxl ou read
    from datetime import datetime

    from django.contrib import messages

    from supplier.models import Supplier
    from supply.models import Item, ItemsCategory

    from .forms import ImportItemsForm

    # Mapping basé sur l'export
    COL_NAME = 'Nom'
    COL_CATEGORY = 'Catégorie'
    COL_AVAILABLE = 'Disponible'
    COL_OUTSIDE = 'Hors site'
    COL_TOTAL = 'Total'
    COL_EXCESS = 'Excédent'
    COL_INFO = 'Informations'
    COL_SUPPLIERS = 'Fournisseurs'
    COL_STOCK_DATE = 'Date entrée stock'
    COL_INV_QTY = 'Qté dernier inventaire'
    COL_INV_DATE = 'Date dernier inventaire'

    if request.method == 'POST':
        form = ImportItemsForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                import_file = request.FILES['import_file']
                update_existing = form.cleaned_data['update_existing']

                data_rows = []

                # --- LECTURE DU FICHIER ---
                if import_file.name.endswith('.csv'):
                    # CSV: Lecture en texte décodé
                    # On utilise io.TextIOWrapper si on veut streamer, ou read().decode()
                    content = import_file.read().decode('utf-8-sig')
                    # On évite les erreurs de lignes vides
                    lines = [l for l in content.splitlines() if l.strip()]
                    reader = csv.DictReader(lines, delimiter=';')
                    data_rows = list(reader)

                elif import_file.name.endswith(('.xlsx', '.xls')):
                    # Excel: Via openpyxl
                    import openpyxl
                    wb = openpyxl.load_workbook(import_file, data_only=True)
                    ws = wb.active

                    headers = [cell.value for cell in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        # Création dict {header: value}
                        row_dict = {}
                        has_data = False
                        for i, cell_val in enumerate(row):
                            if i < len(headers) and headers[i]:
                                row_dict[headers[i]] = cell_val
                                if cell_val:
                                    has_data = True
                        if has_data:
                            data_rows.append(row_dict)

                # --- TRAITEMENT ---
                created_count = 0
                updated_count = 0
                errors = []

                for idx, row in enumerate(data_rows):
                    line_num = idx + 2
                    try:
                        # 1. Nom (Obligatoire)
                        name_val = row.get(COL_NAME)
                        if not name_val:
                            continue  # Ligne vide ou sans nom
                        name = str(name_val).strip()

                        # 2. Catégorie
                        cat_val = row.get(COL_CATEGORY)
                        cat_name = str(cat_val).strip(
                        ) if cat_val else 'Sans catégorie'
                        category, _ = ItemsCategory.objects.get_or_create(
                            name=cat_name)

                        # 3. Recherche Item
                        item = Item.objects.filter(
                            name=name, category=category).first()
                        is_new = False

                        if item:
                            if not update_existing:
                                continue  # On ignore
                        else:
                            item = Item(name=name, category=category)
                            item.created_by = request.user
                            is_new = True

                        # 4. Valeurs numériques
                        def parse_int(val):
                            if val is None or str(val).strip() == '':
                                return 0
                            try:
                                return int(float(str(val).replace(',', '.')))
                            except:
                                return 0

                        item.available_quantity = parse_int(
                            row.get(COL_AVAILABLE))
                        item.outside_quantity = parse_int(row.get(COL_OUTSIDE))
                        item.excess_quantity = parse_int(row.get(COL_EXCESS))
                        item.last_inventory_quantity = parse_int(
                            row.get(COL_INV_QTY))

                        # Total
                        total_val = parse_int(row.get(COL_TOTAL))
                        if total_val > 0:
                            item.total_quantity = total_val
                        else:
                            item.total_quantity = item.available_quantity + item.outside_quantity

                        # Info
                        info_val = row.get(COL_INFO)
                        item.information = str(info_val) if info_val else ''

                        # Date entrée stock (format DD/MM/YYYY attendu du CSV, ou datetime du Excel)
                        stock_date_val = row.get(COL_STOCK_DATE)
                        if stock_date_val:
                            if isinstance(stock_date_val, datetime):
                                item.stock_entry_date = stock_date_val.date()
                            elif isinstance(stock_date_val, str):
                                try:
                                    item.stock_entry_date = datetime.strptime(
                                        stock_date_val, '%d/%m/%Y'
                                    ).date()
                                except:
                                    pass  # Format invalide ignoré

                        item.save()

                        # 5. Fournisseurs
                        sup_val = row.get(COL_SUPPLIERS)
                        if sup_val:
                            # Split virgule
                            sup_names = [s.strip() for s in str(
                                sup_val).split(',') if s.strip()]
                            for s_name in sup_names:
                                # On crée le fournisseur s'il n'existe pas (choix de facilité pour l'import)
                                supplier, _ = Supplier.objects.get_or_create(
                                    name=s_name)
                                item.suppliers.add(supplier)

                        if is_new:
                            created_count += 1
                        else:
                            updated_count += 1

                    except Exception as e_row:
                        errors.append(f"Ligne {line_num} : {e_row}")

                # Feedback
                if errors:
                    msg = f"Import terminé avec {len(errors)} erreurs. (Succès : {created_count} créés, {updated_count} maj)"
                    messages.warning(request, msg)
                    # On pourrait afficher les erreurs plus en détail mais message flash limité
                else:
                    msg = f"Import terminé avec succès ! ({created_count} créés, {updated_count} mis à jour)"
                    messages.success(request, msg)

                return redirect(f"{reverse('supplies_management')}#items")

            except Exception as e:
                messages.error(
                    request, f"Erreur critique lors de l'import : {e}")
                return redirect(f"{reverse('supplies_management')}#items")
    else:
        form = ImportItemsForm()

    return render(request, 'import_items.html', {'form': form})
