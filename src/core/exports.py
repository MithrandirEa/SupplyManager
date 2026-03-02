"""
Fonctions d'export pour la page Matériels.
Formats supportés : CSV (UTF-8 BOM, compatible Google Sheets) et Excel (.xlsx).
"""
import csv
import io
from datetime import datetime

from django.http import HttpResponse


# ─────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────

def _csv_response(filename: str) -> tuple:
    """Retourne un HttpResponse CSV et le writer associé."""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response, delimiter=';')
    return response, writer


def _excel_response(wb, filename: str) -> HttpResponse:
    """Sérialise un workbook openpyxl dans un HttpResponse."""
    from openpyxl.styles import Font, PatternFill, Alignment
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _style_header_row(ws):
    """Met en gras et en couleur la première ligne d'un worksheet."""
    from openpyxl.styles import Font, PatternFill
    fill = PatternFill(start_color='2F4F6F', end_color='2F4F6F', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill


def _autofit(ws):
    """Ajuste approximativement la largeur des colonnes."""
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


# ─────────────────────────────────────────────
#  ITEMS
# ─────────────────────────────────────────────

ITEMS_HEADERS = [
    'Nom', 'Catégorie', 'Disponible', 'Hors site', 'Total',
    'Excédent', 'Informations', 'Fournisseurs',
    'Date entrée stock', 'Qté dernier inventaire', 'Date dernier inventaire',
]


def _items_row(item):
    return [
        item.name,
        item.category.name if item.category else '',
        item.available_quantity,
        item.outside_quantity,
        item.total_quantity,
        item.excess_quantity,
        item.information or '',
        ', '.join(s.name for s in item.suppliers.all()),
        item.stock_entry_date.strftime('%d/%m/%Y') if item.stock_entry_date else '',
        item.last_inventory_quantity,
        item.last_inventory_date.strftime('%d/%m/%Y %H:%M') if item.last_inventory_date else '',
    ]


def export_items_csv(items):
    response, writer = _csv_response('materiels.csv')
    writer.writerow(ITEMS_HEADERS)
    for item in items:
        writer.writerow(_items_row(item))
    return response


def export_items_excel(items):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Matériels'
    ws.append(ITEMS_HEADERS)
    for item in items:
        ws.append(_items_row(item))
    _style_header_row(ws)
    _autofit(ws)
    return _excel_response(wb, 'materiels.xlsx')


# ─────────────────────────────────────────────
#  ORDERS — liste
# ─────────────────────────────────────────────

ORDERS_LIST_HEADERS = [
    '#', 'Fournisseur', 'Date commande', 'Retour attendu',
    'Retour réel', 'Statut', 'Nb articles', 'Notes',
]


def _order_list_row(order):
    return [
        order.id,
        order.supplier.name,
        order.order_date.strftime('%d/%m/%Y %H:%M'),
        order.expected_return_date.strftime('%d/%m/%Y'),
        order.actual_return_date.strftime('%d/%m/%Y') if order.actual_return_date else '',
        order.get_status_display(),
        order.order_items.count(),
        order.notes or '',
    ]


def export_orders_list_csv(orders):
    response, writer = _csv_response('commandes.csv')
    writer.writerow(ORDERS_LIST_HEADERS)
    for order in orders:
        writer.writerow(_order_list_row(order))
    return response


def export_orders_list_excel(orders):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Commandes'
    ws.append(ORDERS_LIST_HEADERS)
    for order in orders:
        ws.append(_order_list_row(order))
    _style_header_row(ws)
    _autofit(ws)
    return _excel_response(wb, 'commandes.xlsx')


# ─────────────────────────────────────────────
#  ORDERS — détail d'une commande
# ─────────────────────────────────────────────

ORDER_DETAIL_ITEM_HEADERS = [
    'Article', 'Catégorie', 'Qté commandée', 'Qté reçue', 'Resté fournisseur',
]


def _order_detail_item_row(oi):
    return [
        oi.item.name,
        oi.item.category.name if oi.item.category else '',
        oi.quantity,
        oi.received_quantity if oi.received_quantity is not None else '',
        oi.remaining_at_supplier if oi.received_quantity is not None else '',
    ]


def export_order_detail_csv(order):
    response, writer = _csv_response(f'commande_{order.id}.csv')
    # En-tête de la commande
    writer.writerow(['Commande #', order.id])
    writer.writerow(['Fournisseur', order.supplier.name])
    writer.writerow(['Date commande', order.order_date.strftime('%d/%m/%Y %H:%M')])
    writer.writerow(['Retour attendu', order.expected_return_date.strftime('%d/%m/%Y')])
    writer.writerow(['Retour réel', order.actual_return_date.strftime('%d/%m/%Y') if order.actual_return_date else ''])
    writer.writerow(['Statut', order.get_status_display()])
    writer.writerow(['Notes', order.notes or ''])
    writer.writerow([])
    writer.writerow(ORDER_DETAIL_ITEM_HEADERS)
    for oi in order.order_items.select_related('item__category').all():
        writer.writerow(_order_detail_item_row(oi))
    return response


def export_order_detail_excel(order):
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Commande {order.id}'

    # Infos générales
    ws.append(['Commande #', order.id])
    ws.append(['Fournisseur', order.supplier.name])
    ws.append(['Date commande', order.order_date.strftime('%d/%m/%Y %H:%M')])
    ws.append(['Retour attendu', order.expected_return_date.strftime('%d/%m/%Y')])
    ws.append(['Retour réel', order.actual_return_date.strftime('%d/%m/%Y') if order.actual_return_date else ''])
    ws.append(['Statut', order.get_status_display()])
    ws.append(['Notes', order.notes or ''])
    ws.append([])

    # En-tête articles
    header_row = ws.max_row + 1
    ws.append(ORDER_DETAIL_ITEM_HEADERS)
    _style_header_row(ws)
    # Re-style only the actual header row
    from openpyxl.styles import Font, PatternFill
    fill = PatternFill(start_color='2F4F6F', end_color='2F4F6F', fill_type='solid')
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
    # Unstyle info rows
    for row in ws.iter_rows(min_row=1, max_row=header_row - 1):
        for cell in row:
            cell.font = Font(bold=False, color='000000')
            cell.fill = PatternFill(fill_type=None)
    # Bold keys in info section
    for row in ws.iter_rows(min_row=1, max_row=header_row - 2, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)

    for oi in order.order_items.select_related('item__category').all():
        ws.append(_order_detail_item_row(oi))
    _autofit(ws)
    return _excel_response(wb, f'commande_{order.id}.xlsx')


# ─────────────────────────────────────────────
#  ORDERS — toutes les commandes avec détail
# ─────────────────────────────────────────────

ORDERS_ALL_DETAIL_HEADERS = [
    'Commande #', 'Fournisseur', 'Date commande', 'Retour attendu',
    'Retour réel', 'Statut', 'Notes commande',
    'Article', 'Catégorie', 'Qté commandée', 'Qté reçue', 'Resté fournisseur',
]


def _orders_all_detail_rows(orders):
    rows = []
    for order in orders:
        for oi in order.order_items.select_related('item__category').all():
            rows.append([
                order.id,
                order.supplier.name,
                order.order_date.strftime('%d/%m/%Y %H:%M'),
                order.expected_return_date.strftime('%d/%m/%Y'),
                order.actual_return_date.strftime('%d/%m/%Y') if order.actual_return_date else '',
                order.get_status_display(),
                order.notes or '',
                oi.item.name,
                oi.item.category.name if oi.item.category else '',
                oi.quantity,
                oi.received_quantity if oi.received_quantity is not None else '',
                oi.remaining_at_supplier if oi.received_quantity is not None else '',
            ])
        if not order.order_items.exists():
            rows.append([
                order.id, order.supplier.name,
                order.order_date.strftime('%d/%m/%Y %H:%M'),
                order.expected_return_date.strftime('%d/%m/%Y'),
                order.actual_return_date.strftime('%d/%m/%Y') if order.actual_return_date else '',
                order.get_status_display(), order.notes or '',
                '', '', '', '', '',
            ])
    return rows


def export_orders_all_detail_csv(orders):
    response, writer = _csv_response('commandes_detail.csv')
    writer.writerow(ORDERS_ALL_DETAIL_HEADERS)
    for row in _orders_all_detail_rows(orders):
        writer.writerow(row)
    return response


def export_orders_all_detail_excel(orders):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Commandes détail'
    ws.append(ORDERS_ALL_DETAIL_HEADERS)
    for row in _orders_all_detail_rows(orders):
        ws.append(row)
    _style_header_row(ws)
    _autofit(ws)
    return _excel_response(wb, 'commandes_detail.xlsx')


# ─────────────────────────────────────────────
#  INVENTORIES — liste
# ─────────────────────────────────────────────

INVENTORIES_LIST_HEADERS = [
    '#', 'Date', 'Créé par', 'Nb articles', 'Notes',
]


def _inventory_list_row(inv):
    return [
        inv.id,
        inv.created_at.strftime('%d/%m/%Y %H:%M'),
        inv.created_by.username if inv.created_by else '',
        inv.entries.count(),
        inv.notes or '',
    ]


def export_inventories_list_csv(inventories):
    response, writer = _csv_response('inventaires.csv')
    writer.writerow(INVENTORIES_LIST_HEADERS)
    for inv in inventories:
        writer.writerow(_inventory_list_row(inv))
    return response


def export_inventories_list_excel(inventories):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventaires'
    ws.append(INVENTORIES_LIST_HEADERS)
    for inv in inventories:
        ws.append(_inventory_list_row(inv))
    _style_header_row(ws)
    _autofit(ws)
    return _excel_response(wb, 'inventaires.xlsx')


# ─────────────────────────────────────────────
#  INVENTORIES — détail d'un inventaire
# ─────────────────────────────────────────────

INVENTORY_DETAIL_HEADERS = [
    'Catégorie', 'Article', 'Sur site', 'Chez fournisseur', 'Total',
]


def _inventory_entry_row(entry):
    return [
        entry.item.category.name if entry.item.category else '',
        entry.item.name,
        entry.counted_quantity,
        entry.outside_quantity_snapshot,
        entry.total_counted,
    ]


def export_inventory_detail_csv(inventory):
    response, writer = _csv_response(f'inventaire_{inventory.id}.csv')
    writer.writerow(['Inventaire #', inventory.id])
    writer.writerow(['Date', inventory.created_at.strftime('%d/%m/%Y %H:%M')])
    writer.writerow(['Créé par', inventory.created_by.username if inventory.created_by else ''])
    writer.writerow(['Notes', inventory.notes or ''])
    writer.writerow([])
    writer.writerow(INVENTORY_DETAIL_HEADERS)
    for entry in inventory.entries.select_related('item__category').order_by('item__category__name', 'item__name'):
        writer.writerow(_inventory_entry_row(entry))
    return response


def export_inventory_detail_excel(inventory):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Inventaire {inventory.id}'

    ws.append(['Inventaire #', inventory.id])
    ws.append(['Date', inventory.created_at.strftime('%d/%m/%Y %H:%M')])
    ws.append(['Créé par', inventory.created_by.username if inventory.created_by else ''])
    ws.append(['Notes', inventory.notes or ''])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(INVENTORY_DETAIL_HEADERS)

    fill = PatternFill(start_color='2F4F6F', end_color='2F4F6F', fill_type='solid')
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
    for row in ws.iter_rows(min_row=1, max_row=header_row - 2, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)

    for entry in inventory.entries.select_related('item__category').order_by('item__category__name', 'item__name'):
        ws.append(_inventory_entry_row(entry))

    _autofit(ws)
    return _excel_response(wb, f'inventaire_{inventory.id}.xlsx')


# ─────────────────────────────────────────────
#  INVENTORIES — tout le détail
# ─────────────────────────────────────────────

INVENTORIES_ALL_DETAIL_HEADERS = [
    'Inventaire #', 'Date', 'Créé par', 'Notes inventaire',
    'Catégorie', 'Article', 'Sur site', 'Chez fournisseur', 'Total',
]


def _inventories_all_detail_rows(inventories):
    rows = []
    for inv in inventories:
        entries = inv.entries.select_related('item__category').order_by('item__category__name', 'item__name')
        for entry in entries:
            rows.append([
                inv.id,
                inv.created_at.strftime('%d/%m/%Y %H:%M'),
                inv.created_by.username if inv.created_by else '',
                inv.notes or '',
                entry.item.category.name if entry.item.category else '',
                entry.item.name,
                entry.counted_quantity,
                entry.outside_quantity_snapshot,
                entry.total_counted,
            ])
        if not entries.exists():
            rows.append([inv.id, inv.created_at.strftime('%d/%m/%Y %H:%M'),
                         inv.created_by.username if inv.created_by else '',
                         inv.notes or '', '', '', '', '', ''])
    return rows


def export_inventories_all_detail_csv(inventories):
    response, writer = _csv_response('inventaires_detail.csv')
    writer.writerow(INVENTORIES_ALL_DETAIL_HEADERS)
    for row in _inventories_all_detail_rows(inventories):
        writer.writerow(row)
    return response


def export_inventories_all_detail_excel(inventories):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventaires détail'
    ws.append(INVENTORIES_ALL_DETAIL_HEADERS)
    for row in _inventories_all_detail_rows(inventories):
        ws.append(row)
    _style_header_row(ws)
    _autofit(ws)
    return _excel_response(wb, 'inventaires_detail.xlsx')


# ─────────────────────────────────────────────
#  MONTHLY STATS
# ─────────────────────────────────────────────

MONTHLY_STATS_HEADERS = [
    'Mois', 'Article', 'Catégorie', 'Quantité envoyée', 'Quantité reçue', 'Quantité facturée',
    'Différence Reçu', 'Différence Facturé'
]


def _get_monthly_stats_rows():
    from collections import defaultdict
    from supplier.models import OrderItem
    from supply.models import Item
    from django.db.models import Sum
    from django.db.models.functions import TruncMonth, Coalesce

    # 1. Tous les items (disponibles)
    items = Item.objects.filter(is_available=True).select_related('category').order_by('category__name', 'name')
    
    # 2. Stats
    raw_stats = OrderItem.objects.annotate(
        month=TruncMonth('order__order_date')
    ).values('month', 'item').annotate(
        sent=Coalesce(Sum('quantity'), 0),
        received=Coalesce(Sum('received_quantity'), 0),
        invoiced=Coalesce(Sum('invoiced_quantity'), 0),
    ).order_by('-month', 'item')

    # 3. Grouper par mois
    grouped_stats = defaultdict(dict)
    months = set()
    
    for stat in raw_stats:
        m = stat['month']
        if m:
            months.add(m)
            grouped_stats[m][stat['item']] = stat

    # On trie les mois du plus récent au plus ancien
    sorted_months = sorted(list(months), reverse=True)
    
    rows = []
    # Pour l'export, on veut peut-être dans l'ordre chronologique ? 
    # Le user voit le plus récent en premier au dashboard, mais un export est souvent chrono.
    # Restons cohérent avec l'UI : décroissant.
    
    for m in sorted_months:
        month_str = m.strftime('%m/%Y')
        for item in items:
            stats = grouped_stats[m].get(item.id, {'sent': 0, 'received': 0, 'invoiced': 0})
            s = stats['sent']
            r = stats['received']
            i = stats['invoiced']
            
            # Si aucune activité pour cet item ce mois-ci, faut-il l'afficher ?
            # L'UI l'affiche (matrice complète). Donc oui.
            
            rows.append([
                month_str,
                item.name,
                item.category.name if item.category else '',
                s,
                r,
                i,
                r - s,      # Diff Reçu (Reçu - Envoyé)
                i - s       # Diff Facturé (Facturé - Envoyé)
            ])
    return rows


def export_monthly_stats_csv():
    response, writer = _csv_response('suivi_mensuel.csv')
    writer.writerow(MONTHLY_STATS_HEADERS)
    for row in _get_monthly_stats_rows():
        writer.writerow(row)
    return response


def export_monthly_stats_excel():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Suivi Mensuel'
    ws.append(MONTHLY_STATS_HEADERS)
    for row in _get_monthly_stats_rows():
        ws.append(row)
    _style_header_row(ws)
    _autofit(ws)
    return _excel_response(wb, 'suivi_mensuel.xlsx')

